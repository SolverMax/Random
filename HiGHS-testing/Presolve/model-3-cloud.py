# Import dependencies
import pyomo.environ as pyo
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# Generic loader from Excel file, given worksheet and named range
def LoadFromExcel(ExcelFile, Worksheet, Range):
    wb = load_workbook(filename=ExcelFile, read_only = True)
    ws = wb[Worksheet]
    dests = wb.defined_names[Range].destinations
    for title, coord in dests:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        data = ws.iter_rows(min_row, max_row, min_col, max_col, values_only = True)
    df = pd.DataFrame(data)
    return df

# Write model to file, if required. The format will be inferred by Pyomo from the file extension, e.g. .gams or .nl
def WriteModelToFile(WriteFile, Model):
    if WriteFile:
        Model.write(ModelFile, io_options={'symbolic_solver_labels': False})   # symbolic_solver_labels of True is easier to read, but a longer file

# Create the Solver object for either NEOS or a local solver
def SetUpSolver(Model):
    Model.Options = None
    if Neos:
        Solver = pyo.SolverManagerFactory('neos')   # Solver on NEOS
        if pyo.value(Model.Engine) == 'cplex':   # Linear
            Model.Options = {'timelimit': Model.TimeLimit}
        elif pyo.value(Model.Engine) == 'octeract':   # Linear or non-linear
            Model.Options = {'MAX_SOLVER_TIME': Model.TimeLimit, 'MILP_SOLVER': 'HIGHS'}
        elif pyo.value(Model.Engine) == 'couenne':   # Non-linear
            print('No options for Couenne')
        else:
            print('Unknown NEOS solver when setting options')
    else:
        Solver = pyo.SolverFactory(pyo.value(Model.Engine))   # Local solver installed
        if pyo.value(Model.Engine) == 'couenne':   # Non-linear
            print('No options for Couenne') # Couenne doesn't accept command line options, use couenne.opt instead
        elif pyo.value(Model.Engine) == 'appsi_highs':   # Linear
            Solver.options['time_limit'] = pyo.value(Model.TimeLimit)
            Solver.options['log_file'] = 'highs.log'   # Sometimes HiGHS doesn't update the console as it solves, so write log file too
            Solver.options['presolve'] = 'on'
        else:
            print('Unknown local solver when setting options')
    
    return Solver, Model

# Call either NEOS or a local solver
def CallSolver(Solver, Model):
    if Neos:
        if Model.Options == None:
            Results = Solver.solve(Model, load_solutions = LoadSolution, tee = Verbose, solver = Model.Engine)
        else:
            Results = Solver.solve(Model, load_solutions = LoadSolution, tee = Verbose, solver = Model.Engine, options = Model.Options)
    else:
        Results = Solver.solve(Model, load_solutions = LoadSolution, tee = Verbose)
    
    return Results, Model

# Load data from Excel file
def GetData(DataFile, DataWorksheet):
    Width = LoadFromExcel(DataFile, DataWorksheet, 'Width')
    Length = LoadFromExcel(DataFile, DataWorksheet, 'Length')
    Weight = LoadFromExcel(DataFile, DataWorksheet, 'Weight')
    Width.columns = ['Item']
    Length.columns = ['Item']
    Weight.columns = ['Item']
    return Width, Length, Weight

# Define model data, assigning all data to the Model
def DefineModelData(Model, Width, Length, Weight):
    Model.Item = pyo.Set(initialize = range(0, len(Width)))
    Model.Candidate = pyo.Set(initialize = range(0, len(Width) * len(Width)))
    Model.Width = pyo.Param(Model.Item, within = pyo.NonNegativeIntegers, mutable = True)
    Model.Length = pyo.Param(Model.Item, within = pyo.NonNegativeIntegers, mutable = True)
    Model.Weight = pyo.Param(Model.Item, within = pyo.NonNegativeReals, mutable = True)
    
    Model.Baseline = 0   # Total weighted area of all items
    for i in Model.Item:
        Model.Width[i] = Width['Item'][i]
        Model.Length[i] = Length['Item'][i]
        Model.Weight[i] = Weight['Item'][i]
        Model.Baseline += Model.Width[i] * Model.Length[i] * Model.Weight[i]
        
    # Define candidate product sizes using width and length of items, taking item width and lengths independently to enumerate all combinations
    Model.CandidateWidth = pyo.Param(Model.Candidate, within = pyo.NonNegativeIntegers, mutable = True)
    Model.CandidateLength = pyo.Param(Model.Candidate, within = pyo.NonNegativeIntegers, mutable = True)
    Model.CandidateArea = pyo.Param(Model.Candidate, within = pyo.NonNegativeIntegers, mutable = True)
    for i in Model.Item:
        for j in Model.Item:
            Model.CandidateWidth[i * len(Width) + j] = Width['Item'][i]
            Model.CandidateLength[i * len(Width) + j] = Length['Item'][j]
            Model.CandidateArea[i * len(Width) + j] = Width['Item'][i] * Length['Item'][j]

# Define model
def DefineModel(Model):
    Model.Select = pyo.Var(Model.Candidate, domain = pyo.Binary)
    Model.Allocation = pyo.Var(Model.Item, Model.Candidate, within = pyo.Binary, initialize = 0)

    def rule_LBWidth(Model, i):   # Width of allocated product must be at least width of each item it is allocated to
        return sum(Model.Allocation[i, c] * Model.CandidateWidth[c] for c in Model.Candidate) >= Model.Width[i]
    Model.MinWidth = pyo.Constraint(Model.Item, rule = rule_LBWidth)

    def rule_LBLength(Model, i):   # Length of allocated product must be at least length of each item it is allocated to
        return sum(Model.Allocation[i, c] * Model.CandidateLength[c] for c in Model.Candidate) >= Model.Length[i]
    Model.MinLength = pyo.Constraint(Model.Item, rule = rule_LBLength)
    
    def rule_count(Model):   # Select the specified number of products that we want to order
        return sum(Model.Select[c] for c in Model.Candidate) == Model.Orders
    Model.NumOrders = pyo.Constraint(rule = rule_count)

    def rule_only(Model, i, c):   # Allocate an item to a candidate only if that candidate is selected
        return Model.Allocation[i, c] <= Model.Select[c]
    Model.SelectedOnly = pyo.Constraint(Model.Item, Model.Candidate, rule = rule_only)
    
    def rule_once(Model, i):   # Each item is allocated to exactly one product
        return sum(Model.Allocation[i, c] for c in Model.Candidate) == 1
    Model.AllocateOnce = pyo.Constraint(Model.Item, rule = rule_once)

    def rule_Obj(Model):   # Minimize waste = Area of allocated product minus area of item, in total for all items
        return sum(sum(Model.Allocation[i, c] * Model.CandidateArea[c] * Model.Weight[i] for c in Model.Candidate) for i in Model.Item) \
               - sum(Model.Width[i] * Model.Length[i] * Model.Weight[i] for i in Model.Item)
    Model.Obj = pyo.Objective(rule = rule_Obj, sense = pyo.minimize)

def WriteOutput(Model, OrderSize, Results):
    Obj = pyo.value(Model.Obj())
    Products = '['
    for c in Model.Candidate:   # Collate list of selected product sizes
        if np.isclose(pyo.value(Model.Select[c]), 1):   # Binary variable = 1, give-or-take small precision error
            Products += str(pyo.value(Model.CandidateWidth[c])).rjust(6) + ' ' + str(pyo.value(Model.CandidateLength[c])).rjust(6) + ' '
    Products += ']'
    print()
    print(f'Order size:   {OrderSize:<,.0f}')
    print(f'Objective:    {Obj:<,.0f} ({Obj / pyo.value(Model.Baseline):.1%} of baseline)')
    print(f'Products:     {Products}\n')

    pd.options.display.float_format = '{:,.0f}'.format
    ItemsAllocated = pd.DataFrame()
    
    SelectedSizes = []   # Dataframe of solution
    for c in Model.Select:
        if np.isclose(pyo.value(Model.Select[c]), 1):
            SelectedSizes.append(str(pyo.value(Model.CandidateWidth[c])) + 'x' + str(pyo.value(Model.CandidateLength[c])))
            for i in Model.Item:
                ItemsAllocated.at[i, c] = pyo.value(Model.Allocation[i, c])
    ItemsAllocated.columns = SelectedSizes

    ItemSizes = []   # Add item sizes to solution dataframe
    for i in Model.Item:
        ItemSizes.append(str(pyo.value(Model.Width[i])) + 'x' + str(pyo.value(Model.Length[i])))
    ItemsAllocated['Item'] = ItemSizes
    pd.set_option('display.max_rows', None)
    print(ItemsAllocated)

def Case(OrderSize, Width, Length, Weight):
    Model = pyo.ConcreteModel(name = ModelName + ', Order size ' + str(OrderSize))
    print(Model.name.strip('\''))
    print('Data file:', DataFile)
    Model.Engine = SolverName
    Model.TimeLimit = TimeLimit
    Solver, Model = SetUpSolver(Model)
    DefineModelData(Model, Width, Length, Weight)
    Model.Orders = OrderSize
    DefineModel(Model)
    WriteModelToFile(WriteFile, Model)
    Results = CallSolver(Solver, Model)
    WriteOutput(Model, OrderSize, Results)

def Main():
#    Timer('Start');
    Width, Length, Weight = GetData(DataFile, DataWorksheet)
#    Timer('Setup');
    for OrderSize in range(ProductsMin, ProductsMax + 1):   # Run multiple product cases, if required
        Case(OrderSize, Width, Length, Weight)
#    Timer('Solved');
#    Timer('Finish');
#    WriteCheckpoints()
    
# Globals

# Data assumptions
ProductsMin = 6   # >= 1
ProductsMax = 6   # <= number of items
DataFile = 'data-60-actual-sample-sorted.xlsx'
DataWorksheet = 'Data'

# Run options
Verbose = True
LoadSolution = True
TimeLimit = 300   # seconds

# Solver options
Neos = False
SolverName = 'appsi_highs'
#os.environ['NEOS_EMAIL'] = 'your-email@company.com'

# Model file
WriteFile = False
ModelFile = 'model-3.gams'

# Fixed
ModelName = 'Paper coverage - Model 3'
#Checkpoints = []   # List of time checkpoints

import psutil
process = psutil.Process()

Main()

MemoryUsage = process.memory_info()
print(MemoryUsage)